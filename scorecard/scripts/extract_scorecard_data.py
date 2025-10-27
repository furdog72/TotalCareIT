#!/usr/bin/env python3
"""
Extract scorecard data from Excel file and output as JSON
"""
import openpyxl
import json
from pathlib import Path

excel_path = "/Users/charles/Downloads/Scorecard 2025.xlsx"

# Load workbook
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Print all sheet names
print("Available sheets:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Try to find the scorecard sheet
scorecard_sheet = None
for sheet_name in wb.sheetnames:
    if 'scorecard' in sheet_name.lower() or 'q4' in sheet_name.lower() or 'q1' in sheet_name.lower():
        scorecard_sheet = wb[sheet_name]
        print(f"\nUsing sheet: {sheet_name}")
        break

if not scorecard_sheet:
    # Use first sheet
    scorecard_sheet = wb[wb.sheetnames[0]]
    print(f"\nUsing first sheet: {wb.sheetnames[0]}")

# Print first 50 rows to understand structure
print("\nFirst 50 rows of data:")
print("-" * 100)
for i, row in enumerate(scorecard_sheet.iter_rows(min_row=1, max_row=50, values_only=True), 1):
    # Filter out completely empty rows
    if any(cell is not None for cell in row):
        print(f"Row {i}: {row}")

wb.close()
