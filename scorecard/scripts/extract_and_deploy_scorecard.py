#!/usr/bin/env python3
"""
Extract Historical Scorecard Data & Deploy to Website

This script:
1. Extracts ALL 2025 historical data from Excel (Q1-Q4, weeks 1-51)
2. Adds week 10/24 API data (HubSpot + Autotask)
3. Saves as JSON on website
4. Builds HTML with retractable quarters
5. Deploys to S3

This makes the website self-contained - no dependency on local Excel file.
"""

import sys
from pathlib import Path
from dotenv import load_dotenv

# Load env BEFORE imports
load_dotenv()

from datetime import datetime, timezone
import json
from openpyxl import load_workbook

EXCEL_FILE = "/Users/charles/Projects/Reference/Scorecard 2025.xlsx"
OUTPUT_JSON = "website/data/scorecard-2025.json"
OUTPUT_HTML = "website/scorecard.html"

WEEK_START = datetime(2025, 10, 18, tzinfo=timezone.utc)
WEEK_END = datetime(2025, 10, 24, 23, 59, 59, tzinfo=timezone.utc)

def extract_all_historical_data():
    """Extract ALL historical data from Scorecard 2025 Excel"""
    print(f"\n📖 Extracting historical data from Excel...")
    print(f"   File: {EXCEL_FILE}")

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Scorecard 2025"]

    # Extract week dates from row 3 (columns 4 onwards)
    print(f"\n   📅 Reading week dates from row 3...")
    weeks = []
    for col in range(4, 100):  # Check up to column 100
        date_val = ws.cell(3, col).value
        if date_val and isinstance(date_val, datetime):
            month_name = date_val.strftime('%B')
            quarter = get_quarter(date_val.month)
            weeks.append({
                'col': col,
                'date': date_val.strftime('%Y-%m-%d'),
                'display': date_val.strftime('%m/%d'),
                'month': month_name,
                'quarter': quarter,
                'year': 2025
            })
        elif col > 60 and not date_val:
            # Stop after we hit empty cells past column 60
            break

    print(f"   ✅ Found {len(weeks)} weeks of historical data")

    # Extract KPI rows (starting from row 4)
    print(f"\n   📊 Reading KPI data...")
    kpis = []

    for row in range(4, 100):  # Check up to row 100
        kpi_name = ws.cell(row, 1).value

        if not kpi_name:
            # Check if we've hit the end of data
            if row > 50:
                break
            continue

        accountable = ws.cell(row, 2).value
        goal = ws.cell(row, 3).value

        # Check if this is a section header
        is_section = is_section_header(ws.cell(row, 1))

        # Get values for all weeks
        week_values = []
        for week_info in weeks:
            col = week_info['col']
            value = ws.cell(row, col).value

            week_values.append({
                'date': week_info['date'],
                'display': week_info['display'],
                'month': week_info['month'],
                'quarter': week_info['quarter'],
                'value': value if value is not None else ''
            })

        kpis.append({
            'kpi': str(kpi_name).strip(),
            'accountable': str(accountable).strip() if accountable else '',
            'goal': goal if goal is not None else '',
            'is_section': is_section,
            'weeks': week_values
        })

    wb.close()

    print(f"   ✅ Extracted {len(kpis)} KPIs")
    return kpis

def get_quarter(month):
    """Get quarter from month number"""
    if month <= 3:
        return 'Q1'
    elif month <= 6:
        return 'Q2'
    elif month <= 9:
        return 'Q3'
    else:
        return 'Q4'

def is_section_header(cell):
    """Check if cell is a section header (bold, merged, etc.)"""
    if cell.font and cell.font.bold:
        return True
    return False

def collect_api_data_for_week_oct24():
    """Collect API data for week 10/24"""
    print(f"\n📊 Collecting API data for week 10/24...")

    api_data = {}

    # HubSpot Data
    try:
        from api.hubspot_service import HubSpotConfig, HubSpotClient, HubSpotReportingService

        config = HubSpotConfig.from_env()
        client = HubSpotClient(config)
        service = HubSpotReportingService(client)

        jason_metrics = service.get_sales_activity_metrics(WEEK_START, WEEK_END, 'jason@totalcareit.com')
        charles_metrics = service.get_sales_activity_metrics(WEEK_START, WEEK_END, 'charles@totalcareit.com')

        api_data.update({
            'Dials': jason_metrics.get('dials', 0),
            'Conversations with DM': jason_metrics.get('conversations_with_dm', 0),
            "FTA's set": jason_metrics.get('ftas_set', 0),
            'FTA Attended': jason_metrics.get('ftas_set', 0),
            'MRR Closed': jason_metrics.get('mrr_closed', 0),
            'W250 Dials': charles_metrics.get('dials', 0),
            'W250 Conversation with a DM': charles_metrics.get('conversations_with_dm', 0),
            "W250 FTA's Set": charles_metrics.get('ftas_set', 0),
        })

        print(f"   ✅ HubSpot: {len(api_data)} metrics collected")
    except Exception as e:
        print(f"   ⚠️  HubSpot error: {e}")

    # Autotask Data
    try:
        import requests
        import os

        username = os.getenv('AUTOTASK_USERNAME')
        secret = os.getenv('AUTOTASK_SECRET')
        integration_code = os.getenv('AUTOTASK_INTEGRATION_CODE')
        zone_url = os.getenv('AUTOTASK_ZONE_URL')
        roc_board_id = int(os.getenv('AUTOTASK_ROC_BOARD_ID'))

        headers = {
            'Content-Type': 'application/json',
            'UserName': username,
            'Secret': secret,
            'ApiIntegrationcode': integration_code
        }

        # Get ROC tickets
        response = requests.post(
            f'{zone_url}/ATServicesRest/V1.0/Tickets/query',
            headers=headers,
            json={
                'filter': [
                    {'field': 'queueID', 'op': 'eq', 'value': roc_board_id},
                    {'field': 'createDate', 'op': 'gte', 'value': WEEK_START.strftime('%Y-%m-%d')},
                    {'field': 'createDate', 'op': 'lte', 'value': WEEK_END.strftime('%Y-%m-%d')}
                ],
                'MaxRecords': 500
            },
            timeout=30
        )

        tickets = response.json().get('items', [])

        # Calculate metrics
        tickets_opened = len(tickets)
        tickets_closed = sum(1 for t in tickets if t.get('completedDate'))

        # Same day close
        same_day = 0
        for t in tickets:
            if t.get('createDate') and t.get('completedDate'):
                create_dt = datetime.fromisoformat(t['createDate'].replace('Z', '+00:00'))
                complete_dt = datetime.fromisoformat(t['completedDate'].replace('Z', '+00:00'))
                if create_dt.date() == complete_dt.date():
                    same_day += 1

        same_day_pct = round((same_day / tickets_opened * 100), 1) if tickets_opened > 0 else 0

        api_data.update({
            'Reactive Tickets Opened': tickets_opened,
            'Reactive Tickets Closed': tickets_closed,
            'Same Day Close %': same_day_pct,
        })

        print(f"   ✅ Autotask: {tickets_opened} tickets, {same_day_pct}% same-day close")
    except Exception as e:
        print(f"   ⚠️  Autotask error: {e}")

    return api_data

def update_week_oct24_with_api_data(kpis, api_data):
    """Update week 10/24 with API-collected data"""
    print(f"\n📝 Updating week 10/24 with API data...")

    target_date = '2025-10-24'
    updates_made = 0

    for kpi in kpis:
        kpi_name = kpi['kpi']

        # Check if we have API data for this KPI
        if kpi_name in api_data:
            # Find the week 10/24 entry
            for week in kpi['weeks']:
                if week['date'] == target_date:
                    week['value'] = api_data[kpi_name]
                    print(f"   ✓ {kpi_name}: {api_data[kpi_name]}")
                    updates_made += 1
                    break

    print(f"   ✅ Updated {updates_made} metrics")
    return kpis

def build_html_with_retractable_quarters(kpis):
    """Build complete HTML with retractable quarters"""
    print(f"\n🏗️  Building HTML with retractable quarters...")

    # Group weeks by quarter
    quarters = {'Q1': [], 'Q2': [], 'Q3': [], 'Q4': []}

    if kpis and kpis[0]['weeks']:
        for week in kpis[0]['weeks']:
            quarter = week['quarter']
            quarters[quarter].append(week)

    for q in ['Q1', 'Q2', 'Q3', 'Q4']:
        print(f"   {q}: {len(quarters[q])} weeks")

    # Get quarter month ranges for labels
    q1_months = set()
    q2_months = set()
    q3_months = set()
    q4_months = set()

    for week in quarters['Q1']:
        q1_months.add(week['month'])
    for week in quarters['Q2']:
        q2_months.add(week['month'])
    for week in quarters['Q3']:
        q3_months.add(week['month'])
    for week in quarters['Q4']:
        q4_months.add(week['month'])

    q1_label = f"Q1 2025 ({', '.join(sorted(q1_months))})" if q1_months else "Q1 2025"
    q2_label = f"Q2 2025 ({', '.join(sorted(q2_months))})" if q2_months else "Q2 2025"
    q3_label = f"Q3 2025 ({', '.join(sorted(q3_months))})" if q3_months else "Q3 2025"
    q4_label = f"Q4 2025 ({', '.join(sorted(q4_months))})" if q4_months else "Q4 2025"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>2025 Scorecard - TotalCare IT</title>
    <link rel="stylesheet" href="css/styles.css">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap" rel="stylesheet">

    <!-- HubSpot Tracking Code -->
    <script type="text/javascript" id="hs-script-loader" async defer src="//js.hs-scripts.com/8752461.js"></script>

    <style>
        * {{ box-sizing: border-box; }}
        body {{
            font-family: 'Inter', 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 0;
            background: #f5f7fa;
        }}
        .scorecard-main {{
            margin-left: 280px;
            padding: 2rem;
            background: #f5f7fa;
            min-height: 100vh;
            width: calc(100% - 280px);
        }}
        @media (max-width: 768px) {{
            .scorecard-main {{
                margin-left: 0;
                width: 100%;
                padding: 1rem;
            }}
        }}
        .header {{
            background: white;
            padding: 20px;
            margin-bottom: 20px;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            margin: 0 0 10px 0;
            font-size: 2rem;
        }}
        .last-updated {{
            color: #7f8c8d;
            font-size: 14px;
        }}
        .container {{
            overflow-x: auto;
            background: white;
            padding: 20px;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        .scorecard-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 11px;
        }}
        .scorecard-table th,
        .scorecard-table td {{
            border: 1px solid #ddd;
            padding: 6px;
            text-align: center;
            white-space: nowrap;
        }}
        .scorecard-table th {{
            background: #34495e;
            color: white;
            position: sticky;
            top: 0;
            z-index: 10;
            font-weight: 600;
        }}
        .kpi-column {{
            text-align: left;
            font-weight: 600;
            background: #ecf0f1;
            position: sticky;
            left: 0;
            z-index: 5;
            min-width: 200px;
        }}
        .accountable-column {{
            background: #ecf0f1;
            position: sticky;
            left: 200px;
            z-index: 5;
            min-width: 100px;
        }}
        .goal-column {{
            background: #fff3cd;
            position: sticky;
            left: 300px;
            z-index: 5;
            min-width: 60px;
        }}
        .section-header {{
            background: #3498db !important;
            color: white !important;
            font-weight: bold;
            text-align: left !important;
            font-size: 12px;
        }}
        .quarter-header {{
            background: #2c3e50;
            color: white;
            cursor: pointer;
            user-select: none;
            font-weight: 600;
        }}
        .quarter-header:hover {{
            background: #34495e;
        }}
        .quarter-collapsed {{
            display: none;
        }}
        .toggle-icon {{
            margin-right: 5px;
            font-weight: bold;
        }}
        .week-header {{
            font-size: 10px;
            min-width: 50px;
        }}
        .data-cell {{
            font-size: 11px;
        }}
        @media (max-width: 768px) {{
            .scorecard-table {{
                font-size: 10px;
            }}
            .kpi-column {{
                min-width: 150px;
            }}
        }}
    </style>
</head>
<body class="dashboard-page">
    <div class="dashboard-container">
        <!-- Sidebar -->
        <aside class="sidebar">
            <div class="sidebar-header">
                <img src="https://totalcareit.com/wp-content/uploads/2024/11/TotalCareIT-Registered-Logo-White-1024x189.png" alt="TotalCare IT Logo" class="sidebar-logo">
            </div>

            <nav class="sidebar-nav">
                <a href="dashboard.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <rect x="3" y="3" width="7" height="7"/>
                        <rect x="14" y="3" width="7" height="7"/>
                        <rect x="14" y="14" width="7" height="7"/>
                        <rect x="3" y="14" width="7" height="7"/>
                    </svg>
                    <span>Dashboard</span>
                </a>
                <a href="sales-report.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <line x1="12" y1="20" x2="12" y2="10"/>
                        <line x1="18" y1="20" x2="18" y2="4"/>
                        <line x1="6" y1="20" x2="6" y2="16"/>
                    </svg>
                    <span>Sales Report</span>
                </a>
                <a href="linkedin-performance.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <path d="M16 8a6 6 0 0 1 6 6v7h-4v-7a2 2 0 0 0-2-2 2 2 0 0 0-2 2v7h-4v-7a6 6 0 0 1 6-6z"/>
                        <rect x="2" y="9" width="4" height="12"/>
                        <circle cx="4" cy="4" r="2"/>
                    </svg>
                    <span>LinkedIn Performance</span>
                </a>
                <a href="scorecard.html" class="nav-item active">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <line x1="18" y1="20" x2="18" y2="10"/>
                        <line x1="12" y1="20" x2="12" y2="4"/>
                        <line x1="6" y1="20" x2="6" y2="14"/>
                    </svg>
                    <span>Scorecard</span>
                </a>
                <a href="quickbooks.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <rect x="2" y="7" width="20" height="14" rx="2" ry="2"/>
                        <path d="M16 21V5a2 2 0 0 0-2-2h-4a2 2 0 0 0-2 2v16"/>
                    </svg>
                    <span>QuickBooks</span>
                </a>
                <a href="prospective-business.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"/>
                        <circle cx="9" cy="7" r="4"/>
                        <path d="M22 21v-2a4 4 0 0 0-3-3.87"/>
                        <path d="M16 3.13a4 4 0 0 1 0 7.75"/>
                    </svg>
                    <span>Prospective Business</span>
                </a>
                <a href="finance.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <line x1="12" y1="1" x2="12" y2="23"/>
                        <path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/>
                    </svg>
                    <span>Finance</span>
                </a>
                <a href="trumethods-qbr.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/>
                        <polyline points="14 2 14 8 20 8"/>
                        <line x1="16" y1="13" x2="8" y2="13"/>
                        <line x1="16" y1="17" x2="8" y2="17"/>
                        <polyline points="10 9 9 9 8 9"/>
                    </svg>
                    <span>TruMethods QBR</span>
                </a>
            </nav>

            <div class="sidebar-footer">
                <p id="user-email"></p>
                <button id="signout-button" class="signout-btn">Sign Out</button>
            </div>
        </aside>

        <!-- Main Content -->
        <main class="scorecard-main">
            <div class="header">
                <h1>TotalCare IT - 2025 Scorecard</h1>
                <div class="last-updated">Last updated: {datetime.now().strftime('%B %d, %Y at %I:%M %p ET')}</div>
            </div>

            <div class="container">
        <table class="scorecard-table">
            <thead>
                <tr>
                    <th rowspan="2" class="kpi-column">KPI</th>
                    <th rowspan="2" class="accountable-column">Accountable</th>
                    <th rowspan="2" class="goal-column">Goal</th>
"""

    # Add quarter headers with correct labels
    quarter_labels = {'Q1': q1_label, 'Q2': q2_label, 'Q3': q3_label, 'Q4': q4_label}
    for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
        week_count = len(quarters[quarter])
        if week_count > 0:
            html += f'                    <th colspan="{week_count}" class="quarter-header" onclick="toggleQuarter(\'{quarter}\')">'
            html += f'<span class="toggle-icon" id="icon-{quarter}">▼</span> {quarter_labels[quarter]}</th>\n'

    html += '                </tr>\n                <tr>\n'

    # Add week headers
    for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
        for week in quarters[quarter]:
            html += f'                    <th class="week-header week-{quarter}">{week["display"]}</th>\n'

    html += '                </tr>\n            </thead>\n            <tbody>\n'

    # Add data rows
    for kpi in kpis:
        if kpi['is_section']:
            colspan = 3 + sum(len(quarters[q]) for q in ['Q1', 'Q2', 'Q3', 'Q4'])
            html += f'                <tr><td colspan="{colspan}" class="section-header">{kpi["kpi"]}</td></tr>\n'
        else:
            html += '                <tr>\n'
            html += f'                    <td class="kpi-column">{kpi["kpi"]}</td>\n'
            html += f'                    <td class="accountable-column">{kpi["accountable"]}</td>\n'
            html += f'                    <td class="goal-column">{kpi["goal"]}</td>\n'

            # Add week data by quarter
            for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
                for i, week in enumerate(kpi['weeks']):
                    if week['quarter'] == quarter:
                        value = week['value']
                        # Format percentages
                        if isinstance(value, (int, float)) and 'percent' in kpi['kpi'].lower() or '%' in str(kpi['goal']):
                            display_value = f"{value}%" if value else ""
                        else:
                            display_value = value if value != '' else ''
                        html += f'                    <td class="data-cell week-{quarter}">{display_value}</td>\n'

            html += '                </tr>\n'

    html += """            </tbody>
        </table>
            </div>
        </main>
    </div>

    <script src="https://alcdn.msauth.net/browser/2.14.2/js/msal-browser.min.js"></script>
    <script src="js/auth.js"></script>
    <script>
        function toggleQuarter(quarter) {
            const cells = document.querySelectorAll('.week-' + quarter);
            const icon = document.getElementById('icon-' + quarter);

            if (!cells.length) return;

            const isCollapsed = cells[0].classList.contains('quarter-collapsed');

            cells.forEach(cell => {
                if (isCollapsed) {
                    cell.classList.remove('quarter-collapsed');
                } else {
                    cell.classList.add('quarter-collapsed');
                }
            });

            icon.textContent = isCollapsed ? '▼' : '▶';
        }

        // Collapse Q1, Q2, Q3 by default (show Q4 - current quarter)
        window.addEventListener('DOMContentLoaded', () => {
            // Collapse past quarters, keep current quarter (Q4) expanded
            ['Q1', 'Q2', 'Q3'].forEach(q => toggleQuarter(q));
        });
    </script>
</body>
</html>"""

    return html

def main():
    """Main execution"""
    print("=" * 80)
    print("Extract & Deploy 2025 Scorecard to Website")
    print("=" * 80)

    # Step 1: Extract all historical data
    kpis = extract_all_historical_data()

    # Step 2: Collect API data for week 10/24
    api_data = collect_api_data_for_week_oct24()

    # Step 3: Update week 10/24 with API data
    kpis = update_week_oct24_with_api_data(kpis, api_data)

    # Step 4: Save JSON
    print(f"\n💾 Saving JSON...")
    output_json_path = Path(OUTPUT_JSON)
    output_json_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_json_path, 'w') as f:
        json.dump(kpis, f, indent=2)
    print(f"   ✅ Saved: {OUTPUT_JSON}")

    # Step 5: Build HTML
    print(f"\n🏗️  Building HTML...")
    html = build_html_with_retractable_quarters(kpis)

    with open(OUTPUT_HTML, 'w') as f:
        f.write(html)
    print(f"   ✅ Saved: {OUTPUT_HTML}")

    # Step 6: Deploy to S3
    print(f"\n🚀 Deploying to S3...")
    import subprocess

    try:
        # Upload JSON
        subprocess.run([
            'aws', 's3', 'cp', OUTPUT_JSON, 's3://totalcareit.ai/data/scorecard-2025.json',
            '--content-type', 'application/json'
        ], check=True)
        print(f"   ✅ Uploaded JSON")

        # Upload HTML
        subprocess.run([
            'aws', 's3', 'cp', OUTPUT_HTML, 's3://totalcareit.ai/scorecard.html',
            '--content-type', 'text/html'
        ], check=True)
        print(f"   ✅ Uploaded HTML")

        # Invalidate CloudFront
        subprocess.run([
            'aws', 'cloudfront', 'create-invalidation',
            '--distribution-id', 'EBUCQMMFVHWED',
            '--paths', '/scorecard.html', '/data/scorecard-2025.json'
        ], check=True)
        print(f"   ✅ CloudFront invalidated")

    except subprocess.CalledProcessError as e:
        print(f"   ⚠️  Deployment error: {e}")
        print(f"\n   Manual deployment commands:")
        print(f"   aws s3 cp {OUTPUT_JSON} s3://totalcareit.ai/data/")
        print(f"   aws s3 cp {OUTPUT_HTML} s3://totalcareit.ai/")
        print(f"   aws cloudfront create-invalidation --distribution-id E39FZHLDTY2FQZ --paths '/scorecard.html' '/data/*'")

    print("\n" + "=" * 80)
    print("✅ COMPLETE!")
    print("=" * 80)
    print(f"\nScorecard available at: https://totalcareit.ai/scorecard.html")
    print(f"JSON data at: https://totalcareit.ai/data/scorecard-2025.json")
    print("\nHistorical data is now stored on the website!")
    print("No dependency on local Excel file.")
    print("=" * 80)

if __name__ == '__main__':
    main()
