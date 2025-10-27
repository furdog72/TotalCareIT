#!/usr/bin/env python3
"""
Build Complete 2025 Scorecard with Retractable Quarters
- Extracts all Q1-Q4 data from Excel (weeks 1-51)
- Adds week 10/24 API data
- Generates HTML with collapsible quarter columns
"""

import sys
from pathlib import Path
from dotenv import load_dotenv

# Load environment BEFORE imports
load_dotenv()

from datetime import datetime
import json
from openpyxl import load_workbook

# Week 4 of October (Oct 24, 2025)
WEEK_OCT_24_DATE = "10/24"
EXCEL_FILE = "/Users/charles/Projects/Reference/Scorecard 2025.xlsx"
OUTPUT_JSON = "website/data/scorecard-2025-complete.json"
OUTPUT_HTML = "website/scorecard.html"

def extract_all_scorecard_data():
    """Extract all 2025 data from Excel file (51 weeks across Q1-Q4)"""
    print(f"\n📖 Reading Excel file: {EXCEL_FILE}")

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active

    # Get week dates from row 3 (columns D onwards = weeks)
    week_dates = []
    for col in range(4, 60):  # Columns D to BG (51 weeks)
        cell_value = ws.cell(3, col).value
        if cell_value:
            week_dates.append({
                'col': col,
                'date': str(cell_value),
                'quarter': get_quarter_from_column(col)
            })

    print(f"   Found {len(week_dates)} weeks of data")

    # Extract KPI rows (rows 4-51)
    scorecard_data = []

    for row in range(4, 52):  # Rows 4-51 contain KPIs
        kpi_name = ws.cell(row, 1).value  # Column A
        accountable = ws.cell(row, 2).value  # Column B
        goal = ws.cell(row, 3).value  # Column C

        if not kpi_name:
            continue

        # Check if this is a section header (bold/merged cell)
        is_section = is_section_header(ws, row)

        # Get all week values
        weeks_data = []
        for week_info in week_dates:
            col = week_info['col']
            value = ws.cell(row, col).value

            weeks_data.append({
                'date': week_info['date'],
                'value': value if value is not None else '',
                'quarter': week_info['quarter']
            })

        scorecard_data.append({
            'row': row,
            'kpi': str(kpi_name).strip() if kpi_name else '',
            'accountable': str(accountable).strip() if accountable else None,
            'goal': goal,
            'is_section': is_section,
            'weeks': weeks_data
        })

    wb.close()
    return scorecard_data

def get_quarter_from_column(col):
    """Determine quarter from column number"""
    # Columns D-P (4-16): Q1 (Jan-Mar, weeks 1-13)
    # Columns Q-AC (17-29): Q2 (Apr-Jun, weeks 14-26)
    # Columns AD-AP (30-42): Q3 (Jul-Sep, weeks 27-39)
    # Columns AQ-BG (43-59): Q4 (Oct-Dec, weeks 40-52)

    if 4 <= col <= 16:
        return 'Q1'
    elif 17 <= col <= 29:
        return 'Q2'
    elif 30 <= col <= 42:
        return 'Q3'
    elif 43 <= 59:
        return 'Q4'
    return 'Q4'

def is_section_header(ws, row):
    """Check if row is a section header (bold text, merged cells, etc.)"""
    cell = ws.cell(row, 1)

    # Check if font is bold
    if cell.font and cell.font.bold:
        return True

    # Check if it's a merged cell
    if ws.merged_cells:
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return True

    return False

def add_api_data_for_week_oct24(scorecard_data):
    """Add API-collected data for week 10/24"""
    print(f"\n📊 Adding API data for week {WEEK_OCT_24_DATE}...")

    # Import here after dotenv is loaded
    from api.hubspot_service import HubSpotConfig, HubSpotClient, HubSpotReportingService
    from api.autotask_service import AutotaskConfig, AutotaskClient, AutotaskReportingService

    # Collect HubSpot data
    try:
        config = HubSpotConfig.from_env()
        client = HubSpotClient(config)
        service = HubSpotReportingService(client)

        WEEK_START = datetime(2025, 10, 18)
        WEEK_END = datetime(2025, 10, 24, 23, 59, 59)

        jason_metrics = service.get_sales_activity_metrics(WEEK_START, WEEK_END, 'jason@totalcareit.com')
        charles_metrics = service.get_sales_activity_metrics(WEEK_START, WEEK_END, 'charles@totalcareit.com')

        api_data = {
            'Dials': jason_metrics.get('dials', 0),
            'Conversation with a DM': jason_metrics.get('conversations_with_dm', 0),
            "FTA's Set": jason_metrics.get('ftas_set', 0),
            'FTA Attended': jason_metrics.get('ftas_set', 0),
            'MRR Closed': jason_metrics.get('mrr_closed', 0),
            'W250 Dials': charles_metrics.get('dials', 0),
            'W250 Conversation with a DM': charles_metrics.get('conversations_with_dm', 0),
            "W250 FTA's Set": charles_metrics.get('ftas_set', 0),
        }

        print(f"   ✅ HubSpot data collected: {len(api_data)} metrics")

    except Exception as e:
        print(f"   ⚠️  HubSpot error: {e}")
        api_data = {}

    # Try Autotask data (may not have enhanced metrics yet)
    try:
        autotask_config = AutotaskConfig.from_env()
        autotask_client = AutotaskClient(autotask_config)
        autotask_service = AutotaskReportingService(autotask_client)

        tickets = autotask_service.get_ticket_metrics(WEEK_START, WEEK_END)

        # Map basic metrics (we know we have 97 tickets from tests)
        api_data.update({
            'Reactive Tickets Opened': tickets.get('total', 0),
            'Reactive Tickets Closed': len([t for t in tickets.get('tickets', []) if t.get('completedDate')]),
        })

        print(f"   ✅ Autotask data collected")

    except Exception as e:
        print(f"   ⚠️  Autotask error: {e}")

    # Update scorecard data with API values
    for kpi in scorecard_data:
        kpi_name = kpi['kpi']
        if kpi_name in api_data:
            # Find week 10/24 and update value
            for week in kpi['weeks']:
                if week['date'] == WEEK_OCT_24_DATE:
                    week['value'] = api_data[kpi_name]
                    print(f"   📝 Updated '{kpi_name}': {api_data[kpi_name]}")

    return scorecard_data

def build_html_with_retractable_quarters(scorecard_data):
    """Build HTML with collapsible quarter columns"""
    print(f"\n🏗️  Building HTML with retractable quarters...")

    # Group weeks by quarter
    quarters = {'Q1': [], 'Q2': [], 'Q3': [], 'Q4': []}

    if scorecard_data and scorecard_data[0]['weeks']:
        for week in scorecard_data[0]['weeks']:
            quarter = week['quarter']
            if quarter in quarters:
                quarters[quarter].append(week)

    print(f"   Q1: {len(quarters['Q1'])} weeks")
    print(f"   Q2: {len(quarters['Q2'])} weeks")
    print(f"   Q3: {len(quarters['Q3'])} weeks")
    print(f"   Q4: {len(quarters['Q4'])} weeks")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>TotalCare IT - 2025 Scorecard</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: #f5f5f5;
        }}
        .container {{
            max-width: 100%;
            overflow-x: auto;
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #2c3e50;
            margin-bottom: 20px;
        }}
        .scorecard-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 12px;
        }}
        .scorecard-table th,
        .scorecard-table td {{
            border: 1px solid #ddd;
            padding: 8px;
            text-align: center;
        }}
        .scorecard-table th {{
            background: #34495e;
            color: white;
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        .kpi-name {{
            text-align: left;
            font-weight: bold;
            background: #ecf0f1;
            position: sticky;
            left: 0;
            z-index: 5;
        }}
        .section-header {{
            background: #3498db !important;
            color: white !important;
            font-weight: bold;
            text-align: left !important;
        }}
        .quarter-header {{
            background: #2c3e50;
            color: white;
            cursor: pointer;
            user-select: none;
        }}
        .quarter-header:hover {{
            background: #34495e;
        }}
        .quarter-collapsed {{
            display: none;
        }}
        .toggle-icon {{
            margin-right: 5px;
        }}
        .goal-column {{
            background: #fff3cd;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>TotalCare IT - 2025 Scorecard</h1>
        <p>Last updated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>

        <table class="scorecard-table">
            <thead>
                <tr>
                    <th rowspan="2" class="kpi-name">KPI</th>
                    <th rowspan="2">Accountable</th>
                    <th rowspan="2" class="goal-column">Goal</th>
"""

    # Add quarter headers
    for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
        week_count = len(quarters[quarter])
        if week_count > 0:
            html += f'                    <th colspan="{week_count}" class="quarter-header" onclick="toggleQuarter(\'{quarter}\')">'
            html += f'<span class="toggle-icon" id="icon-{quarter}">▼</span>{quarter} 2025 ({week_count} weeks)</th>\n'

    html += '                </tr>\n                <tr>\n'

    # Add week date headers
    for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
        for week in quarters[quarter]:
            html += f'                    <th class="week-{quarter}">{week["date"]}</th>\n'

    html += '                </tr>\n            </thead>\n            <tbody>\n'

    # Add data rows
    for kpi in scorecard_data:
        if kpi['is_section']:
            html += f'                <tr><td colspan="999" class="section-header">{kpi["kpi"]}</td></tr>\n'
        else:
            html += '                <tr>\n'
            html += f'                    <td class="kpi-name">{kpi["kpi"]}</td>\n'
            html += f'                    <td>{kpi["accountable"] or ""}</td>\n'
            html += f'                    <td class="goal-column">{kpi["goal"] or ""}</td>\n'

            # Add week values by quarter
            for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
                for week in kpi['weeks']:
                    if week['quarter'] == quarter:
                        value = week['value']
                        html += f'                    <td class="week-{quarter}">{value}</td>\n'

            html += '                </tr>\n'

    html += """            </tbody>
        </table>
    </div>

    <script>
        function toggleQuarter(quarter) {
            const cells = document.querySelectorAll('.week-' + quarter);
            const icon = document.getElementById('icon-' + quarter);
            const isCollapsed = cells[0].classList.contains('quarter-collapsed');

            cells.forEach(cell => {
                if (isCollapsed) {
                    cell.classList.remove('quarter-collapsed');
                    icon.textContent = '▼';
                } else {
                    cell.classList.add('quarter-collapsed');
                    icon.textContent = '▶';
                }
            });
        }

        // Collapse Q1, Q2, Q3 by default
        window.addEventListener('DOMContentLoaded', () => {
            ['Q1', 'Q2', 'Q3'].forEach(q => toggleQuarter(q));
        });
    </script>
</body>
</html>"""

    return html

def main():
    """Main execution"""
    print("=" * 80)
    print("Building Complete 2025 Scorecard")
    print("=" * 80)

    # Step 1: Extract all data from Excel
    scorecard_data = extract_all_scorecard_data()
    print(f"✅ Extracted {len(scorecard_data)} KPIs")

    # Step 2: Add API data for week 10/24
    scorecard_data = add_api_data_for_week_oct24(scorecard_data)

    # Step 3: Save JSON
    output_json_path = Path(OUTPUT_JSON)
    output_json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_json_path, 'w') as f:
        json.dump(scorecard_data, f, indent=2)
    print(f"✅ Saved JSON: {OUTPUT_JSON}")

    # Step 4: Build HTML
    html = build_html_with_retractable_quarters(scorecard_data)

    output_html_path = Path(OUTPUT_HTML)
    with open(output_html_path, 'w') as f:
        f.write(html)
    print(f"✅ Saved HTML: {OUTPUT_HTML}")

    print("\n" + "=" * 80)
    print("✅ COMPLETE!")
    print("=" * 80)
    print(f"\nNext steps:")
    print(f"1. Review: open {OUTPUT_HTML} in browser")
    print(f"2. Deploy: aws s3 cp {OUTPUT_HTML} s3://totalcareit.ai/")
    print(f"3. Invalidate: aws cloudfront create-invalidation --distribution-id E39FZHLDTY2FQZ --paths '/scorecard.html'")
    print("=" * 80)

if __name__ == '__main__':
    main()
