#!/usr/bin/env python3
"""
Simple Excel to HTML Scorecard Converter
Reads the actual Excel structure and creates an HTML table
"""
import openpyxl
from datetime import datetime
from pathlib import Path
import json

# Paths
excel_file = Path("/Users/charles/Projects/TotalCareIT/Reference/Scorecard 2025.xlsx")
output_html = Path("/Users/charles/Projects/TotalCareIT/website/scorecard-excel.html")

# Load workbook
wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb['Scorecard 2025']

print("Reading Excel structure...")

# Find the actual last column with data
last_col = 3
for col in range(4, 100):  # Check first 100 columns
    has_data = False
    for row in range(1, 20):
        if ws.cell(row=row, column=col).value:
            has_data = True
            break
    if has_data:
        last_col = col
    elif col > last_col + 5:  # Stop if 5 consecutive empty columns
        break

print(f"Last data column: {last_col} ({openpyxl.utils.get_column_letter(last_col)})")

# Read week ending dates from row 3
week_dates = {}
for col in range(4, last_col + 1):
    cell = ws.cell(row=3, column=col)
    if cell.value:
        if isinstance(cell.value, datetime):
            date_str = cell.value.strftime('%m/%d')
            week_dates[col] = date_str
        else:
            week_dates[col] = str(cell.value)[:10]
    else:
        week_dates[col] = ""

print(f"Found {len(week_dates)} week columns")

# Identify Q4 October columns (10/03, 10/10, 10/17, 10/24, 10/31)
q4_oct_cols = []
for col, date_str in week_dates.items():
    if date_str.startswith('10/'):  # October dates
        q4_oct_cols.append(col)

print(f"Q4 October columns: {q4_oct_cols}")
print(f"Q4 October dates: {[week_dates[col] for col in q4_oct_cols]}")

# Read all data rows
rows_data = []
for row_idx in range(4, 80):  # Read up to row 80
    metric = ws.cell(row=row_idx, column=1).value
    if not metric:
        continue

    metric_str = str(metric).strip()
    if not metric_str:
        continue

    owner = ws.cell(row=row_idx, column=2).value
    goal = ws.cell(row=row_idx, column=3).value

    # Check if section header (has background color)
    cell_a = ws.cell(row=row_idx, column=1)
    is_section = False
    bg_color = None

    if cell_a.fill and cell_a.fill.start_color:
        color = cell_a.fill.start_color
        if hasattr(color, 'rgb') and color.rgb:
            rgb = str(color.rgb)
            if rgb != '00000000' and len(rgb) >= 6:
                is_section = True
                bg_color = f"#{rgb[-6:]}"

    # Read week values (only Q4 October for now)
    week_values = {}
    for col in q4_oct_cols:
        cell_value = ws.cell(row=row_idx, column=col).value
        week_values[col] = cell_value

    rows_data.append({
        'row': row_idx,
        'metric': metric_str,
        'owner': str(owner) if owner else '',
        'goal': str(goal) if goal else '',
        'is_section': is_section,
        'bg_color': bg_color,
        'weeks': week_values
    })

print(f"Extracted {len(rows_data)} rows")

# Generate HTML
html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>TotalCareIT Scorecard 2025 - Q4</title>
    <link rel="stylesheet" href="css/styles.css">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap" rel="stylesheet">
    <script src="https://alcdn.msauth.net/browser/2.30.0/js/msal-browser.min.js"></script>
    <style>
        body {{ margin: 0; padding: 0; font-family: 'Inter', sans-serif; }}
        .scorecard-wrapper {{
            margin: 20px;
            background: white;
            border-radius: 12px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.08);
            overflow: hidden;
        }}
        .scorecard-header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 2rem;
            text-align: center;
        }}
        .scorecard-header h1 {{
            margin: 0 0 0.5rem 0;
            font-size: 2rem;
        }}
        .table-container {{
            overflow: auto;
            max-height: calc(100vh - 250px);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }}
        thead {{
            position: sticky;
            top: 0;
            z-index: 10;
            background: white;
        }}
        th {{
            padding: 10px 8px;
            background: #34495e;
            color: white;
            border: 1px solid #d0d0d0;
            font-weight: 600;
            white-space: nowrap;
        }}
        th.metric-col {{
            min-width: 250px;
            text-align: left;
            background: #2c3e50;
            position: sticky;
            left: 0;
            z-index: 11;
        }}
        th.week-col {{
            min-width: 80px;
            background: #667eea;
        }}
        td {{
            padding: 8px;
            border: 1px solid #e0e0e0;
            text-align: center;
        }}
        td.metric-cell {{
            text-align: left;
            font-weight: 600;
            background: #f8f9fa;
            position: sticky;
            left: 0;
            z-index: 5;
            border-right: 2px solid #ccc;
        }}
        td.section-header {{
            font-weight: 700;
            font-size: 14px;
            text-align: left;
        }}
        .good {{ background: #d4edda !important; color: #155724; }}
        .warning {{ background: #fff3cd !important; color: #856404; }}
        .bad {{ background: #f8d7da !important; color: #721c24; }}
    </style>
</head>
<body class="dashboard-page">
    <div class="dashboard-container">
        <aside class="sidebar">
            <div class="sidebar-header">
                <img src="https://totalcareit.com/wp-content/uploads/2024/11/TotalCareIT-Registered-Logo-Full-Color.webp" alt="Logo" class="sidebar-logo">
            </div>
            <nav class="sidebar-nav">
                <a href="dashboard.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/>
                        <rect x="14" y="14" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/>
                    </svg><span>Dashboard</span>
                </a>
                <a href="sales-report.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <line x1="12" y1="20" x2="12" y2="10"/><line x1="18" y1="20" x2="18" y2="4"/>
                        <line x1="6" y1="20" x2="6" y2="16"/>
                    </svg><span>Sales Report</span>
                </a>
                <a href="linkedin-performance.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <path d="M16 8a6 6 0 0 1 6 6v7h-4v-7a2 2 0 0 0-2-2 2 2 0 0 0-2 2v7h-4v-7a6 6 0 0 1 6-6z"/>
                        <rect x="2" y="9" width="4" height="12"/><circle cx="4" cy="4" r="2"/>
                    </svg><span>LinkedIn Performance</span>
                </a>
                <a href="scorecard-excel.html" class="nav-item active">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <line x1="18" y1="20" x2="18" y2="10"/><line x1="12" y1="20" x2="12" y2="4"/>
                        <line x1="6" y1="20" x2="6" y2="14"/>
                    </svg><span>Scorecard</span>
                </a>
                <a href="contract-monitor.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/>
                        <polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/>
                        <line x1="16" y1="17" x2="8" y2="17"/><line x1="10" y1="9" x2="8" y2="9"/>
                    </svg><span>Contract Monitor</span>
                </a>
                <a href="quickbooks.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <rect x="2" y="7" width="20" height="14" rx="2" ry="2"/>
                        <path d="M16 21V5a2 2 0 0 0-2-2h-4a2 2 0 0 0-2 2v16"/>
                    </svg><span>QuickBooks</span>
                </a>
                <a href="prospective-business.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"/>
                        <circle cx="9" cy="7" r="4"/>
                        <path d="M22 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/>
                    </svg><span>Prospective Business</span>
                </a>
                <a href="finance.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <line x1="12" y1="1" x2="12" y2="23"/>
                        <path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/>
                    </svg><span>Finance</span>
                </a>
                <a href="trumethods-qbr.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/>
                        <polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/>
                        <line x1="16" y1="17" x2="8" y2="17"/><polyline points="10 9 9 9 8 9"/>
                    </svg><span>TruMethods QBR</span>
                </a>
            </nav>
        </aside>

        <main class="dashboard-main">
            <div class="scorecard-wrapper">
                <div class="scorecard-header">
                    <h1>🎯 TotalCareIT Scorecard 2025 - Q4</h1>
                    <p>Q4 Performance Dashboard - October weeks</p>
                </div>

                <div class="table-container">
                    <table>
                        <thead>
                            <tr>
                                <th class="metric-col">Metric / KPI</th>
                                <th>Accountable</th>
                                <th>Week End<br>↓ Goal</th>
"""

# Add week headers
for col in q4_oct_cols:
    html += f'                                <th class="week-col">{week_dates[col]}</th>\n'

html += """                            </tr>
                        </thead>
                        <tbody>
"""

# Add data rows
for row in rows_data:
    if row['is_section']:
        style = f" style='background: {row['bg_color']} !important;'" if row['bg_color'] else ""
        html += f"                            <tr>\n"
        html += f"                                <td class='section-header'{style} colspan='{3 + len(q4_oct_cols)}'>{row['metric']}</td>\n"
        html += f"                            </tr>\n"
    else:
        html += f"                            <tr>\n"
        html += f"                                <td class='metric-cell'>{row['metric']}</td>\n"
        html += f"                                <td>{row['owner']}</td>\n"
        html += f"                                <td>{row['goal']}</td>\n"

        for col in q4_oct_cols:
            value = row['weeks'].get(col, '')
            if value is None or value == '':
                display = '-'
            elif isinstance(value, (int, float)):
                display = f"{value:.0f}" if value == int(value) else f"{value:.2f}"
            else:
                display = str(value)
            html += f"                                <td>{display}</td>\n"

        html += f"                            </tr>\n"

html += """                        </tbody>
                    </table>
                </div>
            </div>
        </main>
    </div>

    <script src="js/auth.js"></script>
</body>
</html>
"""

# Write HTML file
with open(output_html, 'w') as f:
    f.write(html)

print(f"\n✅ HTML generated: {output_html}")
print(f"   Rows: {len(rows_data)}")
print(f"   Q4 October weeks: {len(q4_oct_cols)}")
print(f"\nYou can now view it in a browser!")
