#!/usr/bin/env python3
"""
Convert Scorecard Excel to HTML with frozen headers and collapsible quarters
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Color
import json
from pathlib import Path

# Paths
excel_file = Path("/Users/charles/Projects/TotalCareIT/Reference/Scorecard 2025.xlsx")
output_html = Path("/Users/charles/Projects/TotalCareIT/website/scorecard-excel.html")
output_json = Path("/Users/charles/Projects/TotalCareIT/website/data/scorecard-excel-data.json")

# Load workbook
wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb['Scorecard 2025']  # Main scorecard sheet

# Define column ranges for each quarter based on actual Excel structure
# We need to scan row 3 for dates to find exact columns
# For now, estimate based on weekly structure (52 weeks + summary columns)
quarters = {
    'Q1': list(range(4, 17)),     # Jan-Mar (13 weeks)
    'Q2': list(range(18, 31)),    # Apr-Jun (13 weeks)
    'Q3': list(range(32, 45)),    # Jul-Sep (13 weeks)
    'Q4': list(range(46, 59))     # Oct-Dec (13 weeks)
}

# Will be updated after reading actual dates from Excel
actual_quarters = {}

# Get week headers from row 3
week_headers = []
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=3, column=col)
    week_headers.append(cell.value if cell.value else "")

# Get column headers from row 2
column_headers = []
for col in range(1, 4):  # Metric/KPI, Accountable, Week End Goal
    cell = ws.cell(row=2, column=col)
    column_headers.append(cell.value if cell.value else "")

print(f"Week headers: {len(week_headers)}")
print(f"Max column: {ws.max_column}")
print(f"Max row: {ws.max_row}")
print(f"Q4 columns: {quarters['Q4']}")

# Extract all data starting from row 4
data = []
for row_idx in range(4, min(ws.max_row + 1, 100)):  # Limit to first 100 rows for now
    row_data = {
        'row_num': row_idx,
        'metric': '',
        'owner': '',
        'goal': '',
        'values': {},
        'is_section_header': False,
        'bg_color': None,
        'text_color': None
    }

    # Get metric name (column A)
    metric_cell = ws.cell(row=row_idx, column=1)
    row_data['metric'] = str(metric_cell.value) if metric_cell.value else ''

    # Check if this is a section header (has background color in column A)
    if metric_cell.fill and metric_cell.fill.start_color:
        color = metric_cell.fill.start_color
        if hasattr(color, 'rgb') and color.rgb and str(color.rgb) != '00000000':  # Not black/transparent
            row_data['is_section_header'] = True
            color_str = str(color.rgb)
            if len(color_str) >= 6:
                row_data['bg_color'] = f"#{color_str[-6:]}"

    # Get owner (column B) and goal (column C)
    row_data['owner'] = str(ws.cell(row=row_idx, column=2).value or '')
    row_data['goal'] = str(ws.cell(row=row_idx, column=3).value or '')

    # Get values for all weeks
    for col_idx in range(4, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value
        number_format = cell.number_format if cell else 'General'

        # Store value with its format
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        row_data['values'][col_letter] = {
            'value': value,
            'format': number_format,
            'col_index': col_idx
        }

    # Only add rows that have content
    if row_data['metric'] or any(v['value'] for v in row_data['values'].values()):
        data.append(row_data)

print(f"Extracted {len(data)} rows of data")

# Save data to JSON
output_json.parent.mkdir(parents=True, exist_ok=True)
with open(output_json, 'w') as f:
    json.dump({
        'quarters': quarters,
        'week_headers': week_headers,
        'column_headers': column_headers,
        'data': data
    }, f, indent=2, default=str)

print(f"Data saved to {output_json}")

# Generate HTML
html_template = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>TotalCareIT Scorecard 2025</title>
    <link rel="stylesheet" href="css/styles.css">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap" rel="stylesheet">
    <script src="https://alcdn.msauth.net/browser/2.30.0/js/msal-browser.min.js"></script>
    <style>
        .scorecard-wrapper {
            margin: 20px;
            background: white;
            border-radius: 12px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.08);
            overflow: hidden;
        }

        .scorecard-header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 2rem;
            text-align: center;
        }

        .scorecard-table-container {
            overflow: auto;
            max-height: calc(100vh - 250px);
            position: relative;
        }

        .scorecard-table {
            width: max-content;
            min-width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            font-family: 'Inter', Arial, sans-serif;
        }

        .scorecard-table thead {
            position: sticky;
            top: 0;
            z-index: 100;
            background: white;
        }

        .scorecard-table th {
            padding: 8px 6px;
            text-align: center;
            border: 1px solid #d0d0d0;
            background: #34495e;
            color: white;
            font-weight: 600;
            white-space: nowrap;
            font-size: 12px;
        }

        .scorecard-table th.metric-col {
            min-width: 250px;
            text-align: left;
            background: #2c3e50;
            position: sticky;
            left: 0;
            z-index: 101;
        }

        .scorecard-table th.owner-col {
            min-width: 100px;
            background: #34495e;
        }

        .scorecard-table th.goal-col {
            min-width: 80px;
            background: #34495e;
        }

        .scorecard-table th.quarter-header {
            background: #5a67d8;
            cursor: pointer;
            user-select: none;
            font-weight: 700;
        }

        .scorecard-table th.quarter-header:hover {
            background: #4c51bf;
        }

        .scorecard-table th.quarter-header.expanded {
            background: #48bb78;
        }

        .scorecard-table th.week-header {
            background: #667eea;
            min-width: 70px;
        }

        .scorecard-table td {
            padding: 8px 6px;
            border: 1px solid #e0e0e0;
            white-space: nowrap;
        }

        .scorecard-table td.metric-cell {
            font-weight: 600;
            background: #f8f9fa;
            position: sticky;
            left: 0;
            z-index: 10;
            border-right: 2px solid #ccc;
        }

        .scorecard-table td.section-header {
            font-weight: 700;
            font-size: 14px;
            text-align: left;
            background: #e9ecef !important;
            color: #495057;
        }

        .scorecard-table td.value-cell {
            text-align: center;
            font-weight: 500;
        }

        .scorecard-table .quarter-col {
            display: none;
        }

        .scorecard-table .quarter-col.visible {
            display: table-cell;
        }

        .good { background: #d4edda !important; color: #155724; }
        .warning { background: #fff3cd !important; color: #856404; }
        .bad { background: #f8d7da !important; color: #721c24; }
        .empty { background: #fafafa; color: #999; }
    </style>
</head>
<body class="dashboard-page">
    <div class="dashboard-container">
        <!-- Sidebar (same as other pages) -->
        <aside class="sidebar">
            <div class="sidebar-header">
                <img src="https://totalcareit.com/wp-content/uploads/2024/11/TotalCareIT-Registered-Logo-Full-Color.webp" alt="Logo" class="sidebar-logo">
            </div>
            <nav class="sidebar-nav">
                <a href="dashboard.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/>
                        <rect x="14" y="14" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/>
                    </svg><span>Dashboard</span>
                </a>
                <a href="sales-report.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <line x1="12" y1="20" x2="12" y2="10"/><line x1="18" y1="20" x2="18" y2="4"/>
                        <line x1="6" y1="20" x2="6" y2="16"/>
                    </svg><span>Sales Report</span>
                </a>
                <a href="linkedin-performance.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <path d="M16 8a6 6 0 0 1 6 6v7h-4v-7a2 2 0 0 0-2-2 2 2 0 0 0-2 2v7h-4v-7a6 6 0 0 1 6-6z"/>
                        <rect x="2" y="9" width="4" height="12"/><circle cx="4" cy="4" r="2"/>
                    </svg><span>LinkedIn Performance</span>
                </a>
                <a href="scorecard-excel.html" class="nav-item active">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <line x1="18" y1="20" x2="18" y2="10"/><line x1="12" y1="20" x2="12" y2="4"/>
                        <line x1="6" y1="20" x2="6" y2="14"/>
                    </svg><span>Scorecard</span>
                </a>
                <a href="contract-monitor.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/>
                        <polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/>
                        <line x1="16" y1="17" x2="8" y2="17"/><line x1="10" y1="9" x2="8" y2="9"/>
                    </svg><span>Contract Monitor</span>
                </a>
                <a href="quickbooks.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <rect x="2" y="7" width="20" height="14" rx="2" ry="2"/>
                        <path d="M16 21V5a2 2 0 0 0-2-2h-4a2 2 0 0 0-2 2v16"/>
                    </svg><span>QuickBooks</span>
                </a>
                <a href="prospective-business.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <path d="M16 21v-2a4 4 0 0 0-4-4H6a4 4 0 0 0-4 4v2"/>
                        <circle cx="9" cy="7" r="4"/>
                        <path d="M22 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/>
                    </svg><span>Prospective Business</span>
                </a>
                <a href="finance.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <line x1="12" y1="1" x2="12" y2="23"/>
                        <path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/>
                    </svg><span>Finance</span>
                </a>
                <a href="trumethods-qbr.html" class="nav-item">
                    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                        <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/>
                        <polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/>
                        <line x1="16" y1="17" x2="8" y2="17"/><polyline points="10 9 9 9 8 9"/>
                    </svg><span>TruMethods QBR</span>
                </a>
            </nav>
        </aside>

        <!-- Main Content -->
        <main class="dashboard-main">
            <div class="scorecard-wrapper">
                <div class="scorecard-header">
                    <h1>🎯 TotalCareIT Scorecard 2025</h1>
                    <p>Full Year Performance Dashboard - Click quarters to expand/collapse</p>
                </div>

                <div class="scorecard-table-container">
                    <table class="scorecard-table" id="scorecardTable">
                        <!-- Table will be generated by JavaScript -->
                    </table>
                </div>
            </div>
        </main>
    </div>

    <script>
        // Load data and build table
        fetch('data/scorecard-excel-data.json')
            .then(response => response.json())
            .then(data => {
                buildScorecardTable(data);
            })
            .catch(error => {
                console.error('Error loading scorecard data:', error);
            });

        function buildScorecardTable(data) {
            const table = document.getElementById('scorecardTable');
            const quarters = data.quarters;
            const weekHeaders = data.week_headers;

            // Build header row
            let headerHTML = '<thead><tr>';
            headerHTML += '<th class="metric-col">Metric / KPI</th>';
            headerHTML += '<th class="owner-col">Accountable</th>';
            headerHTML += '<th class="goal-col">Week End<br>↓ Goal</th>';

            // Add quarter headers
            ['Q1', 'Q2', 'Q3', 'Q4'].forEach((q, qIndex) => {
                const colIndices = quarters[q];
                const isExpanded = (q === 'Q4'); // Q4 expanded by default

                // Quarter summary column
                headerHTML += `<th class="quarter-header ${isExpanded ? 'expanded' : ''}" onclick="toggleQuarter('${q}')" id="header-${q}">${q} ▸</th>`;

                // Week columns (hidden for Q1-Q3, visible for Q4)
                colIndices.forEach(colIdx => {
                    const weekHeader = weekHeaders[colIdx - 1] || '';
                    headerHTML += `<th class="week-header quarter-col quarter-${q} ${isExpanded ? 'visible' : ''}">${weekHeader}</th>`;
                });
            });

            headerHTML += '<th class="year-col">Year</th>';
            headerHTML += '</tr></thead>';

            // Build body rows
            let bodyHTML = '<tbody>';
            data.data.forEach(row => {
                if (!row.metric && !row.is_section_header) return; // Skip empty rows

                const isSection = row.is_section_header;
                const bgColor = row.bg_color || '';

                bodyHTML += '<tr>';

                // Metric name
                const metricClass = isSection ? 'section-header' : 'metric-cell';
                const metricStyle = bgColor ? `background: ${bgColor} !important;` : '';
                bodyHTML += `<td class="${metricClass}" style="${metricStyle}">${row.metric}</td>`;

                // Owner
                bodyHTML += `<td>${row.owner}</td>`;

                // Goal
                bodyHTML += `<td>${row.goal}</td>`;

                // Quarter and week values
                ['Q1', 'Q2', 'Q3', 'Q4'].forEach(q => {
                    const isExpanded = (q === 'Q4');

                    // Quarter summary (empty for now - will add later)
                    bodyHTML += `<td class="value-cell"></td>`;

                    // Week values
                    quarters[q].forEach(colIdx => {
                        const colLetter = getColumnLetter(colIdx);
                        const cellData = row.values[colLetter];
                        let cellValue = '';

                        if (cellData && cellData.value !== null && cellData.value !== undefined) {
                            cellValue = formatValue(cellData.value, cellData.format);
                        }

                        bodyHTML += `<td class="value-cell quarter-col quarter-${q} ${isExpanded ? 'visible' : ''}">${cellValue}</td>`;
                    });
                });

                // Year total (empty for now)
                bodyHTML += '<td class="value-cell"></td>';

                bodyHTML += '</tr>';
            });
            bodyHTML += '</tbody>';

            table.innerHTML = headerHTML + bodyHTML;
        }

        function toggleQuarter(quarter) {
            const cols = document.querySelectorAll(`.quarter-${quarter}`);
            const header = document.getElementById(`header-${quarter}`);
            const isVisible = cols[0].classList.contains('visible');

            cols.forEach(col => {
                if (isVisible) {
                    col.classList.remove('visible');
                    header.classList.remove('expanded');
                } else {
                    col.classList.add('visible');
                    header.classList.add('expanded');
                }
            });
        }

        function getColumnLetter(colIndex) {
            let letter = '';
            while (colIndex > 0) {
                const remainder = (colIndex - 1) % 26;
                letter = String.fromCharCode(65 + remainder) + letter;
                colIndex = Math.floor((colIndex - 1) / 26);
            }
            return letter;
        }

        function formatValue(value, format) {
            if (value === null || value === undefined || value === '' || value === '-') {
                return '-';
            }

            const num = parseFloat(value);
            if (isNaN(num)) return value;

            // Handle percentage formats
            if (format && format.includes('%')) {
                const decimals = (format.match(/\\.0+/) || [''])[0].length - 1;
                return (num * 100).toFixed(decimals) + '%';
            }

            // Handle decimal formats
            if (format && format.includes('.')) {
                const decimals = (format.match(/\\.0+/) || [''])[0].length - 1;
                return num.toFixed(decimals);
            }

            // Integer format
            return Math.round(num).toString();
        }
    </script>
</body>
</html>
"""

# Write HTML file
with open(output_html, 'w') as f:
    f.write(html_template)

print(f"HTML generated: {output_html}")
print("\nNext steps:")
print("1. Review the generated HTML file")
print("2. Test the collapsible quarters functionality")
print("3. Add API integration for week ending 10/24")
